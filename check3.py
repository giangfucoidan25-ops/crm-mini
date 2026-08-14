import openpyxl
import glob
file = 'Nhập liệu Quản lý Đơn Hàng - Tháng 5.xlsx'
wb = openpyxl.load_workbook(file, data_only=True)
ws = wb.active
header_row = 1
for r in range(1, 10):
    val3 = str(ws.cell(row=r, column=3).value or '').lower()
    val6 = str(ws.cell(row=r, column=6).value or '').lower()
    if 'điện thoại' in val3 or 'họ tên' in val6:
        header_row = r
        break
print("Header row:", header_row)
header_vals = [str(ws.cell(row=header_row, column=c).value or '').lower() for c in range(1, 15)]
print("Header vals:", header_vals)
