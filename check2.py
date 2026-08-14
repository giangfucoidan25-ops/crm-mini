import pandas as pd
import json
import re
import openpyxl
import glob

files = glob.glob('Nhập liệu Quản lý Đơn Hàng - Tháng 5.xlsx')

phone_to_url = {}
name_to_url = {}

for file in files:
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # find header row
        header_row = 1
        for r in range(1, 10):
            val = ws.cell(row=r, column=3).value or ''
            if 'số điện thoại' in str(val).lower() or 'họ tên' in str(ws.cell(row=r, column=6).value or '').lower():
                header_row = r
                break
                
        for row in ws.iter_rows(min_row=header_row+1):
            header_vals = [str(ws.cell(row=header_row, column=c).value or '').lower() for c in range(1, 15)]
            try:
                phone_col = next(i for i, v in enumerate(header_vals) if 'điện thoại' in v)
                name_col = next(i for i, v in enumerate(header_vals) if 'họ tên' in v)
            except StopIteration:
                phone_col = 4
                name_col = 5

            phone_val = str(row[phone_col].value or '').strip()
            name_val = str(row[name_col].value or '').strip()
            
            if not name_val or name_val == 'nan' or name_val == 'None':
                continue
                
            match = re.search(r'^KH(\d+)', name_val, flags=re.IGNORECASE)
            if match:
                getfly_id = match.group(1)
                
                base_domain = 'fucoidan2026.getflycrm.com'
                if row[name_col].hyperlink:
                    target = row[name_col].hyperlink.target
                    if target and 'hangocchau' in target:
                        base_domain = 'hangocchau.getflycrm.com'
                
                url = f"https://{base_domain}/#/crm/view_account/{getfly_id}"
                
                # clean name to match database
                clean_name = re.sub(r'^KH\d+\s*-\s*', '', name_val, flags=re.IGNORECASE).strip()
                name_to_url[clean_name] = url
                
                phone = ''.join(filter(str.isdigit, phone_val))
                if phone and not phone.startswith('0'): phone = '0' + phone
                if phone:
                    phone_to_url[phone] = url
                    
    except Exception as e:
        print(f"Error processing {file}: {e}")

print("Phones found:", '0937305383' in phone_to_url, '0357847334' in phone_to_url, '0904143336' in phone_to_url)
print("Keys:", list(phone_to_url.keys())[:10])
print("Name found Lê Hiệp:", any('Lê Hiệp' in k for k in name_to_url.keys()))
