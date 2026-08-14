import pandas as pd
import json
import re
import openpyxl

db = json.load(open('database.json', encoding='utf-8'))

import glob
files = glob.glob('Nhập liệu Quản lý Đơn Hàng*.xlsx')

phone_to_url = {}
name_to_url = {}

for file in files:
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # find header row
        header_row = 1
        for r in range(1, 30):
            row_vals = [str(ws.cell(row=r, column=c).value or '').lower() for c in range(1, 15)]
            if any('điện thoại' in v for v in row_vals) or any('họ tên' in v for v in row_vals):
                header_row = r
                break
                
        for row in ws.iter_rows(min_row=header_row+1):
            header_vals = [str(ws.cell(row=header_row, column=c).value or '').lower() for c in range(1, 15)]
            try:
                phone_col = next(i for i, v in enumerate(header_vals) if 'điện thoại' in v)
                name_col = next(i for i, v in enumerate(header_vals) if 'họ tên' in v)
            except StopIteration:
                # default fallback if still not found
                phone_col = 3 # index 3 -> col D
                name_col = 4  # index 4 -> col E

            phone_val = str(row[phone_col].value or '').strip()
            name_val = str(row[name_col].value or '').strip()
            
            if not name_val or name_val == 'nan' or name_val == 'None':
                continue
                
            match = re.search(r'^KH(\d+)', name_val, flags=re.IGNORECASE)
            if match:
                getfly_id = match.group(1)
                
                # Try to get hyperlink
                base_domain = 'fucoidan2026.getflycrm.com'
                if row[name_col].hyperlink:
                    target = row[name_col].hyperlink.target
                    if target and 'hangocchau' in target:
                        base_domain = 'hangocchau.getflycrm.com'
                
                url = f"https://{base_domain}/#/crm/view_account/{getfly_id}"
                
                name_to_url[name_val] = url
                
                # Clean phone
                phone = ''.join(filter(str.isdigit, phone_val))
                if phone and not phone.startswith('0'): phone = '0' + phone
                if phone:
                    phone_to_url[phone] = url
    except Exception as e:
        print(f"Error processing {file}: {e}")

updated = 0
for c in db.get('customers', []):
    if c.get('phone') in phone_to_url:
        c['getfly_url'] = phone_to_url[c['phone']]
        updated += 1
    elif c.get('zalo_phone') in phone_to_url:
        c['getfly_url'] = phone_to_url[c['zalo_phone']]
        updated += 1
    elif not c.get('getfly_url'):
        # match by name if phone is missing or didn't match
        for name_val, url in name_to_url.items():
            if c.get('full_name') and c['full_name'].lower() in name_val.lower():
                c['getfly_url'] = url
                updated += 1
                break

json.dump(db, open('database.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=4)
print(f"Updated {updated} customers with getfly_url.")
